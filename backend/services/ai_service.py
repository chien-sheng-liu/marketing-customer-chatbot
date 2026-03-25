"""AI Service — wraps all OpenAI calls and RAG utilities.

All functions that require an OpenAI client call ensure_client() so that
startup never hard-crashes when the key is missing; errors surface at
request time instead.
"""
from __future__ import annotations

import logging
import math
import os
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List, Optional

from fastapi import HTTPException
from openai import OpenAI
from openpyxl import load_workbook
from pypdf import PdfReader
from docx import Document

from rag_store import RagStore, chunk_text  # noqa: F401 — re-exported for callers

logger = logging.getLogger("marketing_tools.ai_service")

OPENAI_MODEL = os.getenv("OPENAI_MODEL", "gpt-4o-mini")
EMBEDDING_MODEL = os.getenv("OPENAI_EMBEDDING_MODEL", "text-embedding-3-small")

# ---------------------------------------------------------------------------
# Singleton OpenAI client
# ---------------------------------------------------------------------------

_client: Optional[OpenAI] = None


def _build_client() -> Optional[OpenAI]:
    api_key = os.getenv("CHATGPT_API_KEY")
    if not api_key:
        logger.warning("CHATGPT_API_KEY is not set — AI routes will fail until configured.")
        return None
    logger.info("OpenAI client initialised with model %s", OPENAI_MODEL)
    return OpenAI(api_key=api_key)


def ensure_client() -> OpenAI:
    global _client
    if _client is None:
        _client = _build_client()
    if _client is None:
        raise HTTPException(status_code=500, detail="CHATGPT_API_KEY is not configured")
    return _client


# Shared RAG store — initialised once at import time.
rag_store = RagStore()

# ---------------------------------------------------------------------------
# Conversation helpers
# ---------------------------------------------------------------------------


def format_conversation(history: list) -> str:
    if not history:
        return ""
    return "\n".join(
        f"{'Customer' if getattr(msg, 'role', None) == 'user' else 'Agent'}: {msg.content}"
        for msg in history
    )


# ---------------------------------------------------------------------------
# OpenAI completions
# ---------------------------------------------------------------------------


def run_completion(
    system_prompt: str,
    user_prompt: str,
    *,
    temperature: float = 0.2,
    json_mode: bool = False,
) -> str:
    kwargs: dict = {
        "model": OPENAI_MODEL,
        "temperature": temperature,
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ],
    }
    if json_mode:
        kwargs["response_format"] = {"type": "json_object"}

    try:
        response = ensure_client().chat.completions.create(**kwargs)
    except HTTPException:
        raise
    except Exception as error:
        logger.exception("ChatGPT request failed (json_mode=%s)", json_mode)
        raise HTTPException(status_code=502, detail="Failed to contact ChatGPT API") from error

    content = response.choices[0].message.content
    if not content:
        raise HTTPException(status_code=502, detail="No response content from ChatGPT")
    return content


# ---------------------------------------------------------------------------
# Embeddings
# ---------------------------------------------------------------------------


def embed_texts(texts: List[str]) -> List[List[float]]:
    if not texts:
        return []
    try:
        response = ensure_client().embeddings.create(model=EMBEDDING_MODEL, input=texts)
        return [record.embedding for record in response.data]
    except HTTPException:
        raise
    except Exception as error:
        logger.exception("Embedding request failed")
        raise HTTPException(status_code=502, detail="Failed to embed text") from error


# ---------------------------------------------------------------------------
# Cosine similarity + RAG search
# ---------------------------------------------------------------------------


def cosine_similarity(a: List[float], b: List[float]) -> float:
    if not a or not b:
        return 0.0
    dot = sum(x * y for x, y in zip(a, b))
    norm_a = math.sqrt(sum(x * x for x in a))
    norm_b = math.sqrt(sum(y * y for y in b))
    if not norm_a or not norm_b:
        return 0.0
    return dot / (norm_a * norm_b)


def get_rag_matches(query: str, top_k: int) -> List[Dict[str, Any]]:
    query = query.strip()
    if not query:
        return []

    chunk_records = rag_store.load_chunk_records()
    if not chunk_records:
        return []

    query_embedding = embed_texts([query])
    if not query_embedding:
        return []

    scored = []
    for record in chunk_records:
        embedding = record.get("embedding")
        if not embedding:
            continue
        score = cosine_similarity(query_embedding[0], embedding)
        scored.append(
            {
                "docId": record.get("docId"),
                "filename": record.get("originalName"),
                "snippet": record.get("text", "")[:800],
                "score": round(score, 4),
            }
        )

    scored.sort(key=lambda item: item["score"], reverse=True)
    return scored[:top_k]


# ---------------------------------------------------------------------------
# Document text extraction
# ---------------------------------------------------------------------------


def _decode_text(binary: bytes) -> str:
    try:
        return binary.decode("utf-8")
    except UnicodeDecodeError:
        return binary.decode("utf-8", errors="ignore")


def extract_document_text(filename: str, binary: bytes) -> str:
    extension = Path(filename).suffix.lower()
    if extension in {".txt", ".md", ".markdown"}:
        return _decode_text(binary)

    buffer = BytesIO(binary)
    try:
        if extension == ".pdf":
            reader = PdfReader(buffer)
            texts = [(page.extract_text() or "").strip() for page in reader.pages]
            return "\n".join(filter(None, texts)).strip()

        if extension == ".docx":
            document = Document(buffer)
            texts = [p.text.strip() for p in document.paragraphs if p.text.strip()]
            return "\n".join(texts).strip()

        if extension == ".xlsx":
            workbook = load_workbook(buffer, read_only=True, data_only=True)
            rows: List[str] = []
            for sheet in workbook.worksheets:
                rows.append(f"# Sheet: {sheet.title}")
                for row in sheet.iter_rows(values_only=True):
                    values = [str(cell).strip() for cell in row if cell not in (None, "")]
                    if values:
                        rows.append(" | ".join(values))
            workbook.close()
            return "\n".join(rows).strip()
    except Exception as error:
        logger.exception("Failed to parse document %s", filename)
        if extension == ".pdf" and "DependencyError" in str(error):
            raise HTTPException(
                status_code=400,
                detail="此 PDF 使用受支援的加密方式，請解除密碼保護後再上傳",
            ) from error
        raise HTTPException(status_code=400, detail="無法解析上傳的文件內容") from error

    raise HTTPException(status_code=400, detail="不支援的文件格式")
