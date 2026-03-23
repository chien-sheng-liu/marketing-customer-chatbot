import json
import logging
import math
import os
from collections import defaultdict
from datetime import datetime
from io import BytesIO
from pathlib import Path
from threading import Lock
from typing import Dict, List, Optional
from uuid import uuid4

from dotenv import load_dotenv
from fastapi import FastAPI, File, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from openpyxl import load_workbook
from openai import OpenAI
from pypdf import PdfReader
from docx import Document
from pydantic import BaseModel, Field

from knowledge_base import KNOWLEDGE_BASE
from rag_store import RagStore, chunk_text


load_dotenv(dotenv_path=".env.local")
load_dotenv()

LOG_LEVEL = os.getenv("LOG_LEVEL", "INFO").upper()
logging.basicConfig(
    level=getattr(logging, LOG_LEVEL, logging.INFO),
    format="%(asctime)s %(levelname)s [backend] %(message)s",
)
logger = logging.getLogger("marketing_tools.backend")

PORT = int(os.getenv("PORT", "4000"))
OPENAI_MODEL = os.getenv("OPENAI_MODEL", "gpt-4o-mini")
CORS_ORIGINS = os.getenv("CORS_ORIGINS")
EMBEDDING_MODEL = os.getenv("OPENAI_EMBEDDING_MODEL", "text-embedding-3-small")
RAG_CHUNK_SIZE = int(os.getenv("RAG_CHUNK_SIZE", "800"))
RAG_CHUNK_OVERLAP = int(os.getenv("RAG_CHUNK_OVERLAP", "200"))
RAG_MAX_FILE_BYTES = int(os.getenv("RAG_MAX_FILE_BYTES", "2000000"))
RAG_ALLOWED_EXTENSIONS = {'.txt', '.md', '.markdown', '.pdf', '.docx', '.xlsx'}
BASE_DIR = Path(__file__).resolve().parent
RAG_STORAGE_ROOT = Path(os.getenv("RAG_STORAGE_PATH", str(BASE_DIR / "storage")))
DEFAULT_WELCOME_MESSAGE = os.getenv(
    "DEFAULT_WELCOME_MESSAGE",
    "歡迎來到 Kamee Growth Desk。請輸入您的會員編號，讓我們個人化您的服務體驗。"
)


class MessagePayload(BaseModel):
    role: Optional[str] = None
    content: str


class AnalyzeRequest(BaseModel):
    history: List[MessagePayload] = Field(default_factory=list)


class KnowledgeBaseRequest(BaseModel):
    query: str = Field(min_length=1)


class RagSearchRequest(BaseModel):
    query: str = Field(min_length=1)
    topK: int = Field(default=3, ge=1, le=10)


class ConversationMessagePayload(BaseModel):
    role: str
    content: str


def _build_client() -> Optional[OpenAI]:
    api_key = os.getenv("CHATGPT_API_KEY")
    if not api_key:
        logger.warning("CHATGPT_API_KEY is not set. API routes will fail until configured.")
        return None
    logger.info("OpenAI client initialized with model %s", OPENAI_MODEL)
    return OpenAI(api_key=api_key)


client = _build_client()
rag_store = RagStore(RAG_STORAGE_ROOT)
conversation_lock = Lock()
conversations: Dict[str, List[Dict[str, str]]] = defaultdict(list)


def normalize_conversation_id(conversation_id: str) -> str:
    normalized = conversation_id.strip().lower() if conversation_id else 'default'
    return normalized or 'default'

app = FastAPI()

origins = [origin.strip() for origin in (CORS_ORIGINS or "").split(",") if origin.strip()]

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins or ["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


def ensure_client() -> OpenAI:
    if not client:
        raise HTTPException(status_code=500, detail="CHATGPT_API_KEY is not configured")
    return client


def format_conversation(history: List[MessagePayload]) -> str:
    if not history:
        return ""
    return "\n".join(
        f"{'Customer' if msg.role == 'user' else 'Agent'}: {msg.content}"
        for msg in history
    )


def _decode_text(binary: bytes) -> str:
    try:
        return binary.decode('utf-8')
    except UnicodeDecodeError:
        return binary.decode('utf-8', errors='ignore')


def extract_document_text(filename: str, binary: bytes) -> str:
    extension = Path(filename).suffix.lower()
    if extension in {'.txt', '.md', '.markdown'}:
        return _decode_text(binary)

    buffer = BytesIO(binary)

    try:
        if extension == '.pdf':
            reader = PdfReader(buffer)
            texts = [(page.extract_text() or '').strip() for page in reader.pages]
            return '\n'.join(filter(None, texts)).strip()

        if extension == '.docx':
            document = Document(buffer)
            texts = [paragraph.text.strip() for paragraph in document.paragraphs if paragraph.text.strip()]
            return '\n'.join(texts).strip()

        if extension == '.xlsx':
            workbook = load_workbook(buffer, read_only=True, data_only=True)
            rows: List[str] = []
            for sheet in workbook.worksheets:
                rows.append(f"# Sheet: {sheet.title}")
                for row in sheet.iter_rows(values_only=True):
                    values = [str(cell).strip() for cell in row if cell not in (None, '')]
                    if values:
                        rows.append(' | '.join(values))
            workbook.close()
            return '\n'.join(rows).strip()
    except Exception as error:
        logger.exception("Failed to parse document %s", filename)
        if extension == '.pdf' and 'DependencyError' in str(error):
            raise HTTPException(status_code=400, detail="此 PDF 使用受支援的加密方式，請解除密碼保護後再上傳") from error
        raise HTTPException(status_code=400, detail="無法解析上傳的文件內容") from error

    raise HTTPException(status_code=400, detail="不支援的文件格式")


def embed_texts(texts: List[str]) -> List[List[float]]:
    if not texts:
        return []
    try:
        response = ensure_client().embeddings.create(model=EMBEDDING_MODEL, input=texts)
        return [record.embedding for record in response.data]
    except Exception as error:
        logger.exception("Embedding request failed")
        raise HTTPException(status_code=502, detail="Failed to embed text") from error


def cosine_similarity(a: List[float], b: List[float]) -> float:
    if not a or not b:
        return 0.0
    dot_product = sum(x * y for x, y in zip(a, b))
    norm_a = math.sqrt(sum(x * x for x in a))
    norm_b = math.sqrt(sum(y * y for y in b))
    if not norm_a or not norm_b:
        return 0.0
    return dot_product / (norm_a * norm_b)


def get_rag_matches(query: str, top_k: int) -> List[dict]:
    query = query.strip()
    if not query:
        return []

    chunk_records = rag_store.load_chunk_records()
    if not chunk_records:
        return []

    query_embedding = embed_texts([query])
    if not query_embedding:
        return []

    scored = []
    for record in chunk_records:
        embedding = record.get('embedding')
        if not embedding:
            continue
        score = cosine_similarity(query_embedding[0], embedding)
        scored.append({
            'docId': record.get('docId'),
            'filename': record.get('originalName'),
            'snippet': record.get('text', '')[:800],
            'score': round(score, 4),
        })

    scored.sort(key=lambda item: item['score'], reverse=True)
    return scored[:top_k]


def get_conversation_messages(conversation_id: str) -> List[Dict[str, str]]:
    conv_id = normalize_conversation_id(conversation_id)
    with conversation_lock:
        thread_messages = conversations.get(conv_id)
        if not thread_messages:
            message = {
                'id': uuid4().hex,
                'role': 'assistant',
                'content': DEFAULT_WELCOME_MESSAGE,
                'timestamp': datetime.utcnow().isoformat()
            }
            conversations[conv_id] = [message]
            thread_messages = conversations[conv_id]
        return list(thread_messages)


def append_conversation_message(conversation_id: str, role: str, content: str) -> Dict[str, str]:
    conv_id = normalize_conversation_id(conversation_id)
    message = {
        'id': uuid4().hex,
        'role': role,
        'content': content,
        'timestamp': datetime.utcnow().isoformat()
    }
    with conversation_lock:
        conversations[conv_id].append(message)
    return message


def run_completion(system_prompt: str, user_prompt: str, *, temperature: float = 0.2, json_mode: bool = False) -> str:
    kwargs: dict = {
        "model": OPENAI_MODEL,
        "temperature": temperature,
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ],
    }

    if json_mode:
        kwargs["response_format"] = {"type": "json_object"}

    try:
        response = ensure_client().chat.completions.create(**kwargs)
    except Exception as error:
        logger.exception("ChatGPT request failed (json_mode=%s)", json_mode)
        raise HTTPException(status_code=502, detail="Failed to contact ChatGPT API") from error

    content = response.choices[0].message.content
    if not content:
        raise HTTPException(status_code=502, detail="No response content from ChatGPT")
    return content


@app.get("/health")
def health_check():
    return {"status": "ok"}


@app.post("/api/analyze")
def analyze_conversation(payload: AnalyzeRequest):
    conversation_text = format_conversation(payload.history)
    last_user_message = next(
        (msg.content for msg in reversed(payload.history) if (msg.role or '').lower() == 'user'),
        ''
    )
    rag_matches = get_rag_matches(last_user_message, top_k=3) if last_user_message else []
    rag_context = ""
    if rag_matches:
        rag_lines = [
            "ADDITIONAL DOC SNIPPETS (use these for accurate answers):"
        ]
        for idx, match in enumerate(rag_matches, start=1):
            rag_lines.append(
                f"{idx}. Source: {match['filename']} | Score: {match['score']}\n{match['snippet']}"
            )
        rag_context = "\n".join(rag_lines)

    prompt = f"""You are an AI assistant that supports a Taiwanese ecommerce customer-service lead.
Use ONLY the internal knowledge base below to reason about the latest conversation.

INTERNAL KNOWLEDGE BASE:
{KNOWLEDGE_BASE}

{rag_context if rag_context else ''}

CONVERSATION HISTORY:
{conversation_text}

Return a valid JSON object with the following fields:
{{
  \"tags\": array of strings,
  \"sentiment\": one of \"positive\", \"neutral\", \"negative\", \"angry\",
  \"routing\": one of \"none\", \"dietitian\", \"senior_agent\", \"risk_alert\",
  \"suggestedReply\": string written in Traditional Chinese without Markdown,
  \"upsellOpportunity\": {{ \"detected\": boolean, \"suggestion\": string }},
  \"reasoning\": short Traditional Chinese explanation without Markdown
}}

Answer in JSON ONLY."""

    try:
        raw = run_completion(
            "You are an AI copilot for a Taiwanese CS lead. ALWAYS respond with strict JSON and avoid Markdown.",
            prompt,
            json_mode=True,
        )
        return json.loads(raw)
    except json.JSONDecodeError as error:
        logger.exception("Invalid JSON returned by ChatGPT: %s", raw)
        raise HTTPException(status_code=502, detail="Invalid JSON returned by ChatGPT") from error


@app.post("/api/kb-query")
def query_knowledge_base(payload: KnowledgeBaseRequest):
    prompt = f"""You are an internal SOP search bot.
Answer strictly based on the knowledge base below.

INTERNAL SOPs:
{KNOWLEDGE_BASE}

Agent question: {payload.query}

Instructions:
1. Reply in Traditional Chinese using only plain text.
2. If the answer is not covered, be honest about it."""

    result = run_completion(
        "You answer SOP questions in Traditional Chinese plain text.",
        prompt,
    )
    return {"result": result}


@app.post("/api/report")
def generate_daily_report():
    prompt = """請為客服主管生成一份簡潔的「每日營運洞察報告」。
假設今天的趨勢如下：
- 因颱風延遲，物流相關查詢增加 30%。
- 關於「鎂複方 (Magnesium Complex)」庫存的詢問偏多。
- 有 2 起關於「取消訂閱按鈕失效」的憤怒客訴。

格式要求：
1. 使用繁體中文撰寫。
2. 使用純文字，允許以 - 作為項目符號。
3. 結構包含：流量與情緒總覽、主要議題、策略與銷售機會、明日行動建議。
"""

    report = run_completion(
        "You summarize CS insights in Traditional Chinese plain text.",
        prompt,
    )
    return {"report": report}


@app.get("/api/rag/documents")
def list_rag_documents():
    return rag_store.list_documents()


@app.post("/api/rag/documents")
async def upload_rag_document(file: UploadFile = File(...)):
    filename = file.filename or 'document.txt'
    extension = Path(filename).suffix.lower() or '.txt'
    if extension not in RAG_ALLOWED_EXTENSIONS:
        raise HTTPException(status_code=400, detail="Only .txt and .md files are supported")

    binary = await file.read()
    if not binary:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")
    if len(binary) > RAG_MAX_FILE_BYTES:
        raise HTTPException(status_code=400, detail="File exceeds the maximum size limit")

    text_content = extract_document_text(filename, binary)
    if not text_content.strip():
        raise HTTPException(status_code=400, detail="文件中沒有可用的文字內容")
    chunks = chunk_text(text_content, size=RAG_CHUNK_SIZE, overlap=RAG_CHUNK_OVERLAP)
    if not chunks:
        raise HTTPException(status_code=400, detail="Unable to extract text chunks from the file")

    embeddings = embed_texts(chunks)
    summary = rag_store.save_document(
        original_name=filename,
        binary_data=binary,
        chunk_texts=chunks,
        embeddings=embeddings,
        chunk_size=RAG_CHUNK_SIZE,
        chunk_overlap=RAG_CHUNK_OVERLAP,
        embedding_model=EMBEDDING_MODEL,
    )
    return summary


@app.get("/api/rag/documents/{doc_id}/download")
def download_rag_document(doc_id: str):
    document = rag_store.get_document_files(doc_id)
    if not document:
        raise HTTPException(status_code=404, detail="Document not found")
    metadata = document['metadata']
    file_path = document['file_path']
    return FileResponse(file_path, filename=metadata.get('originalName'), media_type='application/octet-stream')


@app.delete("/api/rag/documents/{doc_id}")
def delete_rag_document(doc_id: str):
    if not rag_store.delete_document(doc_id):
        raise HTTPException(status_code=404, detail="Document not found")
    return {"status": "deleted"}


@app.post("/api/rag/search")
def rag_semantic_search(payload: RagSearchRequest):
    query = payload.query.strip()
    if not query:
        raise HTTPException(status_code=400, detail="Query is required")

    matches = get_rag_matches(query, payload.topK)
    return {"matches": matches}


@app.get("/api/conversations/{conversation_id}/messages")
def get_conversation(conversation_id: str):
    conv_id = normalize_conversation_id(conversation_id)
    return {"messages": get_conversation_messages(conv_id)}


@app.post("/api/conversations/{conversation_id}/messages")
def add_conversation_message(conversation_id: str, payload: ConversationMessagePayload):
    conv_id = normalize_conversation_id(conversation_id)
    role = (payload.role or 'user').lower()
    if role not in {"user", "assistant"}:
        raise HTTPException(status_code=400, detail="role 必須為 user 或 assistant")
    message = append_conversation_message(conv_id, role, payload.content)
    return message


if __name__ == "__main__":
    import uvicorn

    uvicorn.run("main:app", host="0.0.0.0", port=PORT, reload=True)
